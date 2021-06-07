import openpyxl


# create workbook
# wb = openpyxl.Workbook()
# for loading the existing workbooks
wb = openpyxl.load_workbook('transactions.xlsx')


# print(wb.sheetnames)

sheet = wb['Sheet1']

# wb.create_sheet('Sheet2',0)
# wb.remove_Sheet(sheet)

# to get the cell
# cell = sheet['a1']
# print(cell.value)
# to change the value

# cell.value = 1
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)


# THE OTHER METHOD IS THE USE THE CELL METHOD
# usefull when iterating over the rows and columns

cell1 = sheet.cell(row=1, column=1)
# print(sheet.max_row)
# print(sheet.max_column)

# for row in range(1, sheet.max_row+1):
#     for column in range(1, sheet.max_column+1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

column = sheet['a']
# print(column)
cells = sheet['a:c']
print(cells)
sheet.append([1, 2, 3])
# sheet.insert_cols
# sheet.insert_rows
# sheet.delete_cols
# sheet.delete_rows
wb.save('transactions2.xlsx')
